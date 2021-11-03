#! python3

from subprocess import Popen
import pprint

from openpyxl.utils import get_column_letter
import pandas as pd

pd.options.mode.chained_assignment = None
pd.set_option('display.max_rows', 30)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 500)


### Useful functions

def df_to_xl(df, path, sheet_name, w={}):
    """
    Given a dataframe df, sheet_name, and weight w, where column A corresponds
    to key=0, e.g., w = {0 : 1.5, 2: 1.3}, puts a weight of 1.5 on column A and
    a weight of 1.3 on column C, then df_to_xl saves the resultant file to path
    where path ends in .xlsx.
    """
    col_size = []
    col_size.append(max(df.index.astype(str).map(len)))
    for col in df.columns:
        m = max(max(df[col].astype(str).map(len)), len(str(col)))
        col_size.append(m)
    for k in w:
        col_size[k] *= w[k]

    with pd.ExcelWriter(path, mode='w', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name)
        ws = writer.sheets[sheet_name]
        for i in range(1, ws.max_column + 1):
            let = get_column_letter(i)
            ws.column_dimensions[let].width = col_size[i - 1]

def highlight_row(row):
    # to do
    pass

class Inventory:
    col_map = {'Display Name' : 'Description',
               'On Hand' : 'Quantity',
               'Item Category' : 'Category',
               'Bin Number' : 'Bin',
               'Inventory Number' : 'Lot Code',
               'Customer Name/Number' : 'Customer SKU'}

    cols = ['Item',
            'Description',
            'Quantity',
            'Category',
            'Customer SKU',
            'Bin',
            'Lot Code']

    def __init__(self, path):
        self.raw = pd.read_csv(path)
        self.raw.rename(columns=self.col_map, inplace=True)

    def show(self):
        pprint.pprint(self.raw)

    def report(self):
        """
        Returns the cleaned dataframe in a "report" form.
        """
        del self.cols[-2:]
        df_raw = self.raw[self.cols]
        df_raw.set_index(self.cols.pop(0), inplace=True)

        dh = df_raw[self.cols[0:1] + self.cols[2:]]
        dh = dh[~dh.index.duplicated(keep='first')]

        dg = df_raw.groupby(df_raw.index)[self.cols[1]].sum()

        df = dh.join(dg)

        return df

    def report_xl(self, path):
        """
        Save the self.report() df to and excel file located at path.
        """
        df_to_xl(self.report(), path, 'Report', {1 : 0.8})
        print('Click: ' + str(path))
        Popen(['open', str(path)])

    def cycle_count(self, path, locations):
        """
        Returns an excel file with quantities only at descired locations.
        locations is a list derived from predetermined locations given in
        the __main__.py file.
        """
        df_raw = self.raw[self.cols]
        df_raw.set_index('Bin', inplace=True)
        df_raw.sort_index(inplace=True)
        df_list = []
        for loc in locations:
            if loc == 'P':
                dg = df_raw.loc[df_raw.index.str.startswith(loc, na=False) &
                    ~df_raw.index.str.startswith('PROD', na=False)]
            else:
                dg = df_raw.loc[df_raw.index.str.startswith(loc, na=False)]
            df_list.append(dg)

        df = pd.concat(df_list)
        pprint.pprint(df)

        df_to_xl(df, path, 'Cycle Count', {0 : 1.8, 1 : 1.8})

        print('Click: ' + str(path))
        Popen(['open', str(path)])
