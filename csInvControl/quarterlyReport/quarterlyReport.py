#! python3

from pathlib import Path
from datetime import date
from subprocess import Popen
import pprint
import os

import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter

from skuMap import sku_map
from classes import Inventory

pd.set_option('display.max_rows', 30)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 500)

today = str(date.today())

dl_dir = Path.home() / Path('Downloads')
proj_dir = Path(__file__).resolve().parents[0]
snap_dir = Path(proj_dir, 'snapshots')

og_xl = snap_dir / Path('CS HB July 12th Inventory w bulk.xlsx')
og_snap = snap_dir / Path('Dean_Inventory_2021-07-06.xlsx')

comp_path = snap_dir / Path('2021_Q2-Q3_Comparison.xlsx')

sc_xl = dl_dir / Path('Products.xlsx')

current_snap = snap_dir / Path('Dean_Inventory_%s.xlsx' % today)

def get_cs_path(dir):
    for folder_name, subfolders, filenames in os.walk(dir):
        for filename in filenames:
            if 'CSSeattleInventory' in filename:
                return dir / Path(filename)

def df_to_xl(df, path, sheet_name, w={}):
    col_size = []
    col_size.append(max(df.index.astype(str).map(len)))
    for col in df.columns:
        m = max(max(df[col].astype(str).map(len)), len(str(col)))
        col_size.append(m)
    for k in w:
        col_size[k] *= w[k]

    with pd.ExcelWriter(path, mode='w', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name)
        ws = writer.sheets[sheet_name]
        for i in range(1, ws.max_column + 1):
            let = get_column_letter(i)
            ws.column_dimensions[let].width = col_size[i - 1]


def get_og_snap(path_in, path_out):
    cols = ['HB SKU',
            'Description',
            'Category',
            'Unit Cost',
            'Quantity',
            'Total Cost']
    col_map = {'Item' : 'HB SKU',
               '032521 Ext Cost' : 'Total Cost',
               '032521 Raw Pricing' : 'Unit Cost',
               'Item Description' : 'Description',
               'Cat.' : 'Category'}

    xl = pd.ExcelFile(path_in)
    df = pd.read_excel(xl, 'Consolidated', header=1)
    df.drop(columns=['Description', 'Category'], inplace=True)
    df.rename(columns=col_map, inplace=True)
    df = df[cols]
    df.set_index(cols[0], inplace=True)

    dg = df.groupby(df.index)[cols[-2:]].sum()

    dh = df[~df.index.duplicated(keep='first')]
    dh = dh[cols[1:4]].join(dg)
    #dh.reset_index(inplace=True)


    d_phi = pd.DataFrame.from_dict(sku_map).T
    d_phi.index.rename('HB SKU', inplace=True)
    dh['Item'] = d_phi['Item']

    dh.reset_index(inplace=True)
    dh.set_index('Item', inplace=True)
    dh.sort_index(inplace=True)
    #d_phi.reset_index(inplace=True)
    #d_phi.rename(columns={'index' : 'HB SKU'})
    #d_phi.set_index('HB SKU', inplace=True)
    #dk = dh.fillna(d_phi['Item'])
    #dk = pd.concat([dh, d_phi])

    #dk = dk.join(pd.DataFrame.from_dict(sku_map).T['Item'])
    #dk = dk.reset_index()
    #dk.set_index('Item', inplace=True)
    #dk.sort_index(inplace=True)

    #pprint.pprint(dh)

    df_to_xl(dh, path_out, 'Initial Inventory', {0 : 1.5})

def get_sc_df(path):
    cats = ['WIP', 'Shipping Supplies']
    col_map = {'Part Number' : 'HB SKU',
               'On Hand Qty' : 'On Hand'}
    cols = ['HB SKU', 'On Hand', 'Category']

    df = pd.read_excel(path, engine='xlrd')
    df.rename(columns=col_map, inplace=True)
    df = df[cols]
    df = df.loc[df[cols[-1]].isin(cats)]
    #pprint.pprint(df)
    return df

def get_ns_df(path):
    col_map = {'Display Name' : 'Description',
               'Item Category' : 'Category',
               'Customer Name/Number' : 'HB SKU'}
    cols = ['Item', 'HB SKU', 'Description', 'Category', 'On Hand']

    inv = Inventory(path)
    df = inv.report()
    df.rename(columns=col_map, inplace=True)
    df = df[cols[1:]]

    dg = pd.DataFrame.from_dict(sku_map).T
    dg.reset_index(inplace=True)
    dg.rename(columns={'index' : 'HB SKU'}, inplace=True)
    dg.set_index('Item', inplace=True)
    #dg = dg['HB SKU']

    dh = df.fillna(dg)

    #pprint.pprint(dh)
    return dh

def get_current_snap(ns_in, sc_in, path):
    df_ns = get_ns_df(ns_in)
    df_sc = get_sc_df(sc_in)

    df = pd.concat([df_ns, df_sc])
    df.index.rename('Item', inplace=True)

    df_to_xl(df, path, 'Snapshot')

    #pprint.pprint(df)

def compare_snaps(old, new, path):
    df_old = pd.read_excel(old, index_col=0)
    df_new = pd.read_excel(new, index_col=0)

    old_col_map = {'Quantity' : '2021 Q2 Qty',
                   'Total Cost' : '2021 Q2 Cost'}
    new_col_map = {'On Hand' : '2021 Q3 Qty'}

    df_old.rename(columns=old_col_map, inplace=True)
    df_new.rename(columns=new_col_map, inplace=True)

    dg = df_old[df_old.columns.tolist()[-3:]]

    df = df_new.join(dg, how='inner')
    df['2021 Q3 Cost'] = df['Unit Cost'] * df['2021 Q3 Qty']

    df['Potential Bad Stock'] = np.where(
            df['2021 Q2 Qty'] == df['2021 Q3 Qty'], 'YES', '')

    cols = ['HB SKU',
            'Description',
            'Category',
            'Unit Cost',
            '2021 Q2 Qty',
            '2021 Q2 Cost',
            '2021 Q3 Qty',
            '2021 Q3 Cost',
            'Potential Bad Stock']
    df = df[cols]

    #pprint.pprint(df_old)
    #pprint.pprint(df_new)
    pprint.pprint(df)

    df_to_xl(df, path, 'Q2 - Q3 Comparison', {0 : 1.5})

def highlight_row(row):
    pass



if __name__ == '__main__':
    #get_og_snap(og_xl, og_snap)
    #Popen(['open', str(og_snap)])

    #get_current_snap(get_cs_path(dl_dir), sc_xl, current_snap)
    Popen(['open', str(current_snap)])

    compare_snaps(og_snap, current_snap, comp_path)
    Popen(['open', str(comp_path)])
