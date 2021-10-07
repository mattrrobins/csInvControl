#! python3

import pandas as pd
import pprint
from subprocess import Popen
from openpyxl.utils import get_column_letter

#pd.options.mode.chained_assignment = None  # default='warn

def df_to_xl(df, path, sheet_name, w={}):
    col_size = []
    col_size.append(max(df.index.astype(str).map(len)))
    for col in df.columns:
        m = max(max(df[col].astype(str).map(len)), len(str(col)))
        col_size.append(m)
    for k in w:
        col_size[k] *= w[k]

    with pd.ExcelWriter(path, mode='w', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name)
        ws = writer.sheets[sheet_name]
        for i in range(1, ws.max_column + 1):
            let = get_column_letter(i)
            ws.column_dimensions[let].width = col_size[i - 1]


class Inventory:
    cols = ['Item',
            'Display Name',
            'On Hand',
            'Item Category',
            'Customer Name/Number',
            'Bin Number',
            'Inventory Number']

    def __init__(self, path):
        self.raw = pd.read_csv(path)

    def show(self):
        pprint.pprint(self.raw)

    def report(self):
        del self.cols[5]
        df_raw = self.raw[self.cols]
        df_raw.set_index(self.cols.pop(0), inplace=True)

        dh = df_raw[self.cols[0:1] + self.cols[2:]]
        dh = dh
        dh = dh[~dh.index.duplicated(keep='first')]

        dg = df_raw.groupby(df_raw.index)[self.cols[1]].sum()

        df = dh.join(dg)

        #df_to_xl(df, path, 'Report', {1 : 0.8})

        #print('Click: ' + str(path))
        #Popen(['open', str(path)])

        return df

    def cycle_count(self, path, locations):
        df_raw = self.raw[self.cols]
        df_raw.set_index('Bin Number', inplace=True)
        df_raw.sort_index(inplace=True)
        df_list = []
        for loc in locations:
            dg = df_raw[df_raw.index.str.startswith(loc, na=False)]
            df_list.append(dg)

        df = pd.concat(df_list)
        pprint.pprint(df)

        col_size = []
        col_size.append(max(df.index.astype(str).map(len)))
        for x in df.columns:
            m = max(max(df[x].astype(str).map(len)), len(str(x)))
            col_size.append(m)
        col_size[0] *= 1.8
        col_size[1] *= 1.8

        with pd.ExcelWriter(path,
                            mode='w',
                            engine='openpyxl') as writer:
            sheet_name = 'Cycle Count'
            df.to_excel(writer, sheet_name)

            ws = writer.sheets[sheet_name]
            for i in range(1, ws.max_column + 1):
                let = get_column_letter(i)
                ws.column_dimensions[let].width = col_size[i - 1]

        print('Click: ' + str(path))
        Popen(['open', str(path)])
